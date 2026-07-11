from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[2]
TEMPLATES = ROOT / "data" / "import-templates"
DEMO = ROOT / "data" / "sample" / "demo-import"
OUTPUT = ROOT / "data" / "sample" / "ai_economy_ledger_sample_import.xlsx"

SHEETS = [
    "companies",
    "metric_definitions",
    "source_registry",
    "source_documents",
    "claims",
    "metric_observations",
]


def parse_csv(path):
    import csv

    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.reader(handle))


def write_sheet(workbook, name, rows):
    sheet = workbook.create_sheet(title=name[:31])
    header_fill = PatternFill("solid", fgColor="1F2937")
    sample_fill = PatternFill("solid", fgColor="FFF7ED")
    note_fill = PatternFill("solid", fgColor="FEF3C7")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    body_font = Font(name="Arial", color="111827")

    sheet.append(["AI Economy Ledger sample import sheet"])
    sheet.append(["All included rows are fictional samples. Do not treat them as verified facts."])
    sheet.append([])

    for row in rows:
        sheet.append(row)

    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True, color="111827")
        cell.fill = note_fill

    for cell in sheet[2]:
        cell.font = Font(name="Arial", italic=True, color="92400E")
        cell.fill = note_fill

    for cell in sheet[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.font = body_font
            cell.fill = sample_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    sheet.freeze_panes = "A5"
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)


def main():
    workbook = Workbook()
    workbook.remove(workbook.active)

    overview = workbook.create_sheet(title="README")
    overview.append(["AI Economy Ledger sample import workbook"])
    overview.append(["Purpose", "Contributor-friendly demo workbook for PR 5 import templates."])
    overview.append(["Safety", "Every row is fictional and must remain is_sample=true."])
    overview.append(["Verification", "Local validation proves sample rows are excluded from verified totals."])
    overview.column_dimensions["A"].width = 18
    overview.column_dimensions["B"].width = 92
    for row in overview.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", bold=cell.column == 1, color="111827")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for sheet_name in SHEETS:
        rows = parse_csv(DEMO / f"{sheet_name}.csv")
        write_sheet(workbook, sheet_name, rows)

    workbook.save(OUTPUT)


if __name__ == "__main__":
    main()
